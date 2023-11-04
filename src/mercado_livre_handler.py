from src.df_handler import DfHandler
from environments import * 
import traceback
from unidecode import unidecode
from openpyxl import load_workbook
from openpyxl.styles import Font
import pandas as pd
from datetime import date
#Data Frame de dados de entrada Mercado Livre
class DFMLInput(DfHandler):
    def __init__(self) -> None:
        super().__init__(file_name=FILE_ML, 
                        sheet_name=SHEET_ANUNCIOS,
                        header_lines=range(0, 4),
                        df=None)
        self.remove_sub_header()
        self.remove_variations()
        self.calc_comissao()
    
    def remove_variations(self):
        self.drop(self[self['VARIATION_ID'].notna()].index, inplace=True)
    
    def calc_comissao(self):

        self["FEE_PER_SALE_MARKETPLACE"] = self.apply(lambda row: 
            "-" if row["CHANNEL"] == "Mercado Shops" else 
            0.115 if row["LISTING_TYPE"] == "Clássico" else
            0.165 if row["LISTING_TYPE"] == "Premium" else "-", axis=1)

#Data Frame de dados de saida Mercado Livre
class DFMLOutput(DfHandler):
    def __init__(self, df=None) -> None:
        if df is not None:
            super().__init__(file_name=None,
                            sheet_name=None,
                            header_lines=None,
                            df=df)
        else:            
            super().__init__(file_name=FILE_CONTROLE_PRECOS, 
                            sheet_name=SHEET_CONTROLE_PRECOS,
                            header_lines=range(0, 1),
                            df=None)
            self.remove_duplicates()

    def add_new_line(self, data_dict):
        new_line = DFMLOutput(data_dict)
        self = DFMLOutput(pd.concat([self, new_line], ignore_index=True))
    
    def remove_duplicates(self):
        self.drop_duplicates(subset='Código ML', inplace=True)


class MLHandler():
    df_o = None
    df_i = None
    def __init__(self, df_input, df_output) -> None:
        self.df_i = df_input
        self.df_o = df_output

    def update_output(self):
        for index, row in self.df_i.iterrows():
            item_id = row['ITEM_ID']
            matching_rows = self.df_o[self.df_o['Código ML'] == item_id]
            if len(matching_rows) == 1:
                matching_row_index  = matching_rows.index[0]   
                self.update_row(matching_row_index, row)

            elif len(matching_rows) < 1:
                self.add_row(row)
                
            else:
                try:
                    raise Exception("Existe um id duplicado na tabela de saída")
                except:
                    print(traceback.print_exc(limit=1))
            
    def update_row(self, index, row):
        self.df_o.at[index, 'SKU'] = row['SKU']
        self.df_o.at[index, 'Título'] = row['TITLE']
        self.df_o.at[index, 'Status'] = row['STATUS']
        self.df_o.at[index, 'Taxa de Comissão (%)'] = row['FEE_PER_SALE_MARKETPLACE']
        preco = row['MARKETPLACE_PRICE']     
        self.df_o.at[index, 'Taxa de Comissão Fixa'] = self.corrige_taxa_fixa(preco) 
        self.df_o.at[index, 'Frete Incluso'] = self.corrige_frete_incluso(row['SHIPPING_METHOD_MARKETPLACE'])
        self.df_o.at[index, 'Preço sem promoção'] = preco
        self.df_o.at[index, 'Data Atualização'] = str(date.today().strftime('%d/%m/%Y'))

    def corrige_taxa_fixa(self, preco):
        """
        Responsável por determinar se tem taxa fixa.
        Retorno: 0.00 ou 5.00
        """
        if preco <= PRECO_TAXA_FIXA:
            return 5.00
        else:
            return 0.00
    
    def add_row(self, row) -> None:
        new_line = pd.DataFrame({'Código ML': [row['ITEM_ID']]})
        self.df_o = DFMLOutput(pd.concat([self.df_o, new_line], copy=False, ignore_index=True))
        new_index = len(self.df_o) - 1        
        self.update_row(new_index, row)
    def corrige_frete_incluso(self, tipo_entrega):
        """
        Responsável por determinar se o frete está incluso ou não.
        Retorno: 'Sim' ou 'Não'
        """
        try:
            tipo_entrega = unidecode(tipo_entrega.strip()).upper()
        except:
            print(traceback.print_exc(limit=1))

        if tipo_entrega in ['MERCADO ENVIOS GRATIS']:
            return 'Sim'
        elif tipo_entrega in ['MERCADO ENVIOS POR CONTA DO COMPRADOR', 'MERCADO ENVIOS POR MI CUENTA']:
            return 'Não'
        else:
            'Erro'
    
    def load_output_to_workbook(self):
        wb = load_workbook(FILE_CONTROLE_PRECOS)
        sheet = wb.active

        for c, (label, content)  in enumerate(self.df_o.items(), 1):
            for r, value in content.items():
                cell = sheet.cell(row=r+2, column=c)
                cell.value = value
                cell.font = Font(size=14)
                if label == 'Taxa de Comissão (%)':
                    cell.number_format = '0.00%'

        wb.save(FILE_CONTROLE_PRECOS)




        print(f'Arquivo {FILE_CONTROLE_PRECOS} atualizado com sucesso!')

    # def format_table(self, sheet):
    #     target_column_name = 'Data Atualização'
    #     target_column_index = self.df_o.columns.get_loc(target_column_name) + 1

    #     for cell in sheet.iter_cols(min_col=target_column_index, max_col=target_column_index):
    #         for c in cell:
    #             c.number_format = 'DD/MM/YYYY'
    #     return sheet











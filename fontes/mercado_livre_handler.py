from fontes.df_handler import DfHandler
from environments import * 
import traceback
from unidecode import unidecode
from openpyxl import load_workbook 
from openpyxl.utils.dataframe import dataframe_to_rows


#Data Frame de dados de entrada Mercado Livre
class DFMLInput(DfHandler):
    def __init__(self) -> None:
        super().__init__(file_name=TABELA_ML, 
                        sheet_name=SHEET_ANUNCIOS,
                        header_lines=range(0, 4))
        self.remove_sub_header()
        self.remove_variations()
        self.calc_comissao()
    
    def remove_variations(self):
        self.drop(self[self['VARIATION_ID'].notna()].index, inplace=True)
    
    def calc_comissao(self):

        self["FEE_PER_SALE_MARKETPLACE"] = self.apply(lambda row: 
            "-" if row["CHANNEL"] == "Mercado Shops" else 
            0.115 if row["LISTING_TYPE"] == "Clássico" else
            0.165 if row["LISTING_TYPE"] == "Premium" else "-", axis=1)

#Data Frame de dados de saida Mercado Livre
class DFMLOutput(DfHandler):
    def __init__(self) -> None:
        super().__init__(file_name=TABELA_CONTROLE_PRECOS, 
                        sheet_name=SHEET_CONTROLE_PRECOS,
                        header_lines=range(0, 1))
        self.drop_duplicates()
    
    def remove_duplicates(self):
        self.drop_duplicates(subset='Código ML', inplace=True)


class MLHandler():
    df_o = None
    df_i = None
    def __init__(self, df_input, df_output) -> None:
        self.df_i = df_input
        self.df_o = df_output

    def update_output(self):
        for index, row in self.df_i.iterrows():
            item_id = row['ITEM_ID']
            matching_rows = self.df_o[self.df_o['Código ML'] == item_id]
            if len(matching_rows) == 1:
                matching_row_index  = matching_rows.index[0]   
                self.update_row(matching_row_index, row, self.df_o)

            elif len(matching_rows) < 1:
                # TODO: Adiciona linha nova na tabela
                pass
            else:
                # TODO: Exclui linha repetida 
                pass
        self.df_o.display_df()

    def update_row(self, index, row, df):
        df.at[index, 'SKU'] = row['SKU']
        df.at[index, 'Título'] = row['TITLE']
        df.at[index, 'Status'] = row['STATUS']
        df.at[index, 'Taxa de Comissão (%)'] = row['FEE_PER_SALE_MARKETPLACE']
        preco = row['MARKETPLACE_PRICE']     
        df.at[index, 'Taxa de Comissão Fixa'] = self.corrige_taxa_fixa(preco) 

        df.at[index, 'Frete Incluso'] = self.corrige_frete_incluso(row['SHIPPING_METHOD_MARKETPLACE'])
        df.at[index, 'Preço sem promoção'] = preco

    def corrige_taxa_fixa(self, preco):
        """
        Responsável por determinar se tem taxa fixa.
        Retorno: 0.00 ou 5.00
        """
        if preco <= PRECO_TAXA_FIXA:
            return 5.00
        else:
            return 0.00
    
    def corrige_frete_incluso(self, tipo_entrega):
        """
        Responsável por determinar se o frete está incluso ou não.
        Retorno: 'Sim' ou 'Não'
        """
        try:
            tipo_entrega = unidecode(tipo_entrega.strip()).upper()
        except:
            print(traceback.print_exc(limit=1))

        if tipo_entrega in ['MERCADO ENVIOS GRATIS']:
            return 'Sim'
        elif tipo_entrega in ['MERCADO ENVIOS POR CONTA DO COMPRADOR', 'MERCADO ENVIOS POR MI CUENTA']:
            return 'Não'
        else:
            'Erro'
    
    def load_output_to_workbook(self):
        # df_values is the DataFrame with the data to save
        wb = load_workbook(TABELA_CONTROLE_PRECOS)
        sheet = wb.active
        df_columns = self.df_o.shape[1]
        # I use 2 with enumerate to save the header column of the excel sheet
        
        for r, row in enumerate(dataframe_to_rows(self.df_o, index=False, header=False), 2):
            for c in range(0, df_columns):
                sheet.cell(row=r, column=c + 1).value = row[c]
        wb.save(TABELA_CONTROLE_PRECOS)











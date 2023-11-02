# Puxar informações via api do bling 
# Atualizar planilha de saida com dados
# Puxar informações da planilha do mercado livre
# Atualizar planilha de saida com os dados

import pandas as pd
from environments import * 
from unidecode import unidecode
from openpyxl import load_workbook 
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import traceback
from fontes.mercado_livre_handler import DfMlInput, DfMlOutput, MLHandler


def main():
    update_dados_planilha()


def update_dados_planilha():
    df_input = DfMlInput()
    df_output = DfMlOutput()
    
    df_input.remove_sub_header()
    df_input.remove_variations()
    
    df_output.display_df()

    for index, row in df_input.iterrows():

        item_id = row['ITEM_ID']
        matching_rows = df_output[df_output['Código ML'] == item_id]
        if len(matching_rows) == 1:
            matching_row_index  = matching_rows.index[0]   
            update_row(matching_row_index, row, df_output)
        
        elif len(matching_rows) < 1:
            # TODO: Adiciona linha nova na tabela
            pass
        else:
            # TODO: Exclui linha repetida 
            pass
    print(df_output.to_string)

    
    # df_values is the DataFrame with the data to save
    wb = load_workbook(TABELA_CONTROLE_PRECOS)
    sheet = wb.active
    df_columns = df_output.shape[1]
    # I use 2 with enumerate to save the header column of the excel sheet
    for r, row in enumerate(dataframe_to_rows(df_output, index=False, header=False), 2):
        for c in range(0, df_columns):
            sheet.cell(row=r, column=c + 1).value = row[c]
    wb.save(TABELA_CONTROLE_PRECOS)


def update_row(index, row, df):
    df.at[index, 'SKU'] = row['SKU']
    df.at[index, 'Título'] = row['TITLE']
    df.at[index, 'Status'] = row['STATUS']
    df.at[index, 'Taxa de Comissão (%)'] = row['FEE_PER_SALE_MARKETPLACE']
    preco = row['MARKETPLACE_PRICE']     
    df.at[index, 'Taxa de Comissão Fixa'] = corrige_taxa_fixa(preco) 

    df.at[index, 'Frete Incluso'] = corrige_frete_incluso(row['SHIPPING_METHOD_MARKETPLACE'])
    df.at[index, 'Preço sem promoção'] = preco


def corrige_taxa_fixa(preco):
    """
    Responsável por determinar se tem taxa fixa.
    Retorno: 0.00 ou 5.00
    """
    if preco <= PRECO_TAXA_FIXA:
        return 5.00
    else:
        return 0.00


def corrige_frete_incluso(tipo_entrega):
    """
    Responsável por determinar se o frete está incluso ou não.
    Retorno: 'Sim' ou 'Não'
    """
    try:
        tipo_entrega = unidecode(tipo_entrega.strip()).upper()
    except:
        print(traceback.print_exc(limit=1))
        

    if tipo_entrega in ['MERCADO ENVIOS GRATIS']:
        return 'Sim'
    elif tipo_entrega in ['MERCADO ENVIOS POR CONTA DO COMPRADOR', 'MERCADO ENVIOS POR MI CUENTA']:
        return 'Não'
    else:
        'Erro'


if __name__ == '__main__':
    main()